from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


def exporta_dfs_to_excel(
    dfs: list[pd.DataFrame],
    template_path: str | Path,
    output_path: str | Path | None = None,
) -> Path:
    template_path = Path(template_path)
    output_path = Path(output_path) if output_path else template_path

    wb = load_workbook(template_path)

    for df in dfs:
        sheet_name = str(df.name)[:31]

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)

        # Encabezados
        for ci, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=ci, value=str(col_name))

        # Datos
        for ri, (_, row) in enumerate(df.iterrows(), start=2):
            for ci, value in enumerate(row, start=1):
                if hasattr(value, "item"):
                    value = value.item()
                ws.cell(row=ri, column=ci, value=value)

    wb.save(output_path)
    return output_path